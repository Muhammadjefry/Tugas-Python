import customtkinter
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import database
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fpdf import FPDF
from datetime import datetime
import csv
import time
import serial
import threading  
import uuid
app = customtkinter.CTk()
app.title("RESISTIVITY MONITORING")
app.geometry("860x480")
app.configure(bg='#161C25')
app.resizable(False, False)

font1 = ("Arial", 20, "bold")
font2 = ("Arial", 12, "bold")
font3 = ("Arial", 12, "normal")


# Inisialisasi Serial
ser = serial.Serial('COM6', 9600, timeout=1)
no_counter = 1
counters = {
    "Sheet1": 1,
    "Sheet2": 1,
    "Sheet3": 1
}
no_counter_lock = threading.Lock()  # Lock untuk menghindari kondisi balapan

def get_active_tab():
    # Ambil nama tab yang aktif dari notebook
    return notebook.tab(notebook.select(), "text")

def start_measurement():
    active_tab = get_active_tab()  # Ambil tab aktif
    count = target_input.get()
    clear_data()  # Hapus hasil sebelumnya

    with no_counter_lock:
        global no_counter
        # Cek jika counter untuk tab aktif belum ada, inisialisasi dengan 1
        if active_tab not in counters:
            counters[active_tab] = 1
        no_counter = counters[active_tab]  # Gunakan counter dari tab aktif

    threading.Thread(target=serial_loop, args=(count,), daemon=True).start()

def auto_measurement():
    start_measurement()  # Panggil fungsi untuk memulai ulang pengukuran otomatis

def manual_measurement():
        global no_counter
        active_tab = get_active_tab()  # Ambil nama tab yang aktif

        # Cek jika counter untuk tab aktif belum ada, inisialisasi dengan 1
        if active_tab not in counters:
            counters[active_tab] = 1

        target_value = target_input.get()

        with no_counter_lock:
            no_counter = counters[active_tab]

            # Cek apakah TreeView sudah penuh atau memenuhi target
            tree = get_active_treeview()  # Mengambil Treeview dari tab aktif
            if len(tree.get_children()) >= int(target_value):
                # Reset counter dan hapus data jika sudah sesuai target
                counters[active_tab] = 1
                clear_data()
                return

            if no_counter > int(target_value):
                counters[active_tab] = 1  # Reset counter untuk tab aktif
                manual_button.configure(state='normal')
                clear_data()
                return

            # Ubah perhitungan volt dan ampere agar mulai dari 0 dan 1
            volt = (no_counter - 1) * 2
            ampere = (no_counter - 1) * 2 + 1
            time_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            unique_id = str(uuid.uuid4())
            short_id = unique_id.replace("-", "")[:8] 


            data = f"{volt},{ampere}"
            ser.write(data.encode())
            print(f"Data sent: {data}")

            tree.insert("", END, values=(short_id, volt, ampere, time_stamp))
            database.insert_measurements(short_id, volt, ampere, time_stamp)
            # Update counter untuk tab aktif
            counters[active_tab] += 1    


# Input counter untuk Automatic dan Manual mode
current_input = IntVar(value=1)
target_input = IntVar(value=10)  # Default target untuk Automatic loop


def get_active_treeview():
    selected_tab = notebook.select()
    tree = notebook.nametowidget(selected_tab).winfo_children()[0]
    return tree

def on_closing():
    if ser.is_open:
        ser.close()
    app.destroy()

app.protocol("WM_DELETE_WINDOW", on_closing)

# Fungsi untuk menambahkan dan menghapus data dari Treeview
def serial_loop(count):
    global no_counter
    tree = get_active_treeview()  # Mengambil Treeview dari tab aktif
    
    for i in range(count):
        volt = i * 2
        ampere = i * 2 + 1
        time_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        unique_id = str(uuid.uuid4())
        short_id = unique_id.replace("-", "")[:8] 

        data = f"{volt},{ampere}"
        ser.write(data.encode())
        print(f"Data sent: {data}")

        tree.insert("", END, values=(short_id, volt, ampere, time_stamp))
        database.insert_measurements(short_id, volt, ampere, time_stamp)

        no_counter += 1
        time.sleep(1)


def clear_data():
    tree = get_active_treeview()  # Mengambil Treeview dari tab aktif
    for item in tree.get_children():
        tree.delete(item)



def update_target_input(step):
    # Update nilai target_input langsung dari tombol Next/Prev
    new_value = target_input.get() + step
    if new_value >= 1:
        target_input.set(new_value)

def add_to_treeview():
    measurements = database.fetch_measurements()
    # tree = notebook.nametowidget(notebook.tabs()[0]).winfo_children()[0]  # Mengakses Treeview dalam tab pertama
    # tree.delete(*tree.get_children())  # Hapus semua data dalam Treeview

    for measurement in measurements:
        tree.insert("", END, values=measurement)

def clear(*clicked):
    if clicked:
        notebook.selection_remove(notebook.focus())
        notebook.focus('')
    id_entry.delete(0, END)
    volt_entry.delete(0, END)
    ampere_entry.delete(0, END)

def clear(*clicked):
    if clicked:
        tree.selection_remove(tree.focus())
        tree.focus('')
    id_entry.delete(0, END)
    volt_entry.delete(0, END)
    ampere_entry.delete(0, END)

# # Fungsi untuk mengambil data dari row yang dipilih dan menampilkannya di Entry
def display_data(event):
    tree = get_active_treeview()
    selected_item = tree.focus()
    if selected_item:
        row = tree.item(selected_item)["values"]
        clear()
        id_entry.insert(0, row[0])
        volt_entry.insert(0, row[1])
        ampere_entry.insert(0, row[2])
    else:
        pass

def delete():
    tree = get_active_treeview()
    selected_item = tree.focus()
    if not selected_item:
        messagebox.showerror("Error", "Choose an measurement to delete")
    else:
        id_val = id_entry.get()
        if id_val:
            try:
                database.delete_measurements(id_val)  # Pastikan fungsi ini benar
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete data: {e}")
                return
        
        tree.delete(selected_item)
        clear()
        messagebox.showinfo("Success", "Data has been deleted")

# Fungsi untuk update data yang dipilih
def update():
    tree = get_active_treeview()
    selected_item = tree.focus()
    if not selected_item:
        messagebox.showerror("Error", "Pilih data untuk diupdate")
    else:
        id_val = id_entry.get()
        volt = volt_entry.get()
        ampere = ampere_entry.get()
        time_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Perbarui data di Treeview
        tree.item(selected_item, values=(id_val, volt, ampere, time_stamp))
        try:
            database.update_measurements(volt, ampere, time_stamp, id_val)
            clear()  # Bersihkan entry setelah update
            messagebox.showinfo("Success", "Data has been updated")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update data: {e}")

        
def export_to_excel():
    # Fetch data from database (Assume measurements is fetched)
    measurements = database.fetch_measurements()
    df = pd.DataFrame(measurements, columns=['ID', 'Volt', 'Ampere', 'Time'])
    
    # Define file name
    file_name = f'measurement_{datetime.now().strftime("%Y-%m-%d_%H.%M.%S")}.xlsx'
    
    # Export DataFrame to Excel
    df.to_excel(file_name, index=False, engine='openpyxl')

    # Load the workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Set title font and style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Apply styles to the header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    column_widths = {
        'A': 10,  # 'No' column width
        'B': 15,  # 'Volt' column width
        'C': 15,  # 'Ampere' column width
        'D': 20   # 'Time' column width
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save the modified workbook
    wb.save(file_name)
    print(f"Data exported to {file_name}")
    messagebox.showinfo("Export Successful", f"Data exported to {file_name}")

def export_to_pdf():
    measurements = database.fetch_measurements()
    info_measurements = database.fetch_measurementsInfo()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=8)
    
    # Menambahkan tabel pertama
    for measurementInfo in info_measurements:  # Gunakan data statis untuk setiap tabel seperti gambar
        pdf.cell(30, 10, 'Operator :', 1)
        pdf.cell(80, 10, str(measurementInfo[0]), 1)
        pdf.cell(30, 10, 'Asisten :', 1)
        pdf.cell(50, 10, str(measurementInfo[1]), 1)
        pdf.ln()

        pdf.cell(30, 10, 'Lokasi :', 1)
        pdf.cell(80, 10, str(measurementInfo[2]), 1)
        pdf.cell(30, 10, 'Tanggal :', 1)
        pdf.cell(50, 10, str(measurementInfo[3]), 1)
        pdf.ln()

        pdf.cell(30, 10, 'Lintasan :', 1)
        pdf.cell(80, 10, str(measurementInfo[4]), 1)
        pdf.cell(30, 10, 'Waktu :', 1)
        pdf.cell(50, 10, str(measurementInfo[5]), 1)
        pdf.ln()

        pdf.cell(30, 10, 'Panjang Lintasan :', 1)
        pdf.cell(15, 10, str(measurementInfo[6]), 1)
        pdf.cell(15, 10, 'Spasi :', 1)
        pdf.cell(10, 10, str(measurementInfo[7]), 1)
        pdf.cell(25, 10, 'Jumlah Channel :', 1)
        pdf.cell(15, 10, str(measurementInfo[8]), 1)
        pdf.cell(30, 10, 'Cuaca :', 1)
        pdf.cell(50, 10, str(measurementInfo[10]), 1)
        pdf.ln()

        pdf.cell(30, 10, 'Koordinat Awal :', 1)
        pdf.cell(80, 10, str(measurementInfo[9]), 1)
        pdf.cell(30, 30, 'Ket :', 1)
        pdf.multi_cell(50, 30, str(measurementInfo[12]), 1)
        pdf.ln()
        
        pdf.cell(30, 10, 'Koordinat Akhir :', 1)
        pdf.cell(80, 10, str(measurementInfo[11]), 1)
        pdf.cell(30, 0, '', 0)  # Ruang kosong agar keterangan tidak pecah
        pdf.cell(50, 0, '', 0)
        pdf.ln()

        pdf.cell(30, 10, 'Elevasi :', 1)
        pdf.cell(80, 10, str(measurementInfo[13]), 1)
        pdf.cell(30, 0, '', 0)  # Ruang kosong agar keterangan tidak pecah
        pdf.cell(50, 0, '', 0)
        pdf.ln(15)
        
      

    # Menambahkan tabel kedua
    pdf.cell(200, 10, txt=f"Measurement Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
    
    # Menambahkan header
    pdf.cell(30, 10, 'ID', 1)
    pdf.cell(50, 10, 'Volt', 1)
    pdf.cell(40, 10, 'Ampere', 1)
    pdf.cell(70, 10, 'Time', 1)
    pdf.ln()

    # Menambahkan data dari measurements
    for measurement in measurements:
        pdf.cell(30, 10, str(measurement[0]), 1)  # Konversi ID ke string
        pdf.cell(50, 10, str(measurement[1]), 1)  # Konversi Volt ke string
        pdf.cell(40, 10, str(measurement[2]), 1)  # Konversi Ampere ke string
        pdf.cell(70, 10, str(measurement[3]), 1)  # Konversi Time ke string
        pdf.ln()    

    # Simpan file PDF
    file_name = f'measurements_{datetime.now().strftime("%Y-%m-%d_%H.%M.%S")}.pdf'
    pdf.output(file_name)
    print(f"Data exported to {file_name}")
    messagebox.showinfo("Export Successful", f"Data exported to {file_name}")

# Fungsi untuk menampilkan waktu sekarang
def show_current_time():
    current_time = datetime.now().strftime("%Y-%m-%d (%H:%M:%S)")
    time_label.configure(text=f"{current_time}")
    app.after(1000, show_current_time)


def export_to_csv():  
    measurements = database.fetch_measurements()
    with open('measurement.csv', mode='w', newline='') as file:  
        writer = csv.writer(file)  
        writer.writerow(['ID', 'Volt', 'Ampere', 'Time'])  # Tulis header CSV  
        writer.writerows(measurements)  # Tulis data karyawan ke CSV
    messagebox.showinfo("Export Successful", "Data exported to measurement.csv")
# Tambahkan label untuk menampilkan jam dan tanggal



# Fungsi untuk membuka jendela baru dengan layout tabel
def open_info_measurement():
    # Buat jendela baru
    
    new_window = customtkinter.CTkToplevel(app)
    new_window.title("Info Measurement")
    new_window.geometry("860x320")

    customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
    customtkinter.set_default_color_theme("dark-blue")  

    def save_data():
        # Mengambil data dari entry
        data = {
            "operator": operator_entry.get(),
            "lokasi": lokasi_entry.get(),
            "lintasan": lintasan_entry.get(),
            "panjang_lintasan": pjglintasan_entry.get(),
            "spasi": spasi_entry.get(),
            "jumlah_chanel": jmlchanel_entry.get(),
            "kordinat_awal": korawal_entry.get(),
            "kordinat_akhir": korakhir_entry.get(),
            "elevasi": elevasi_entry.get(),
            "asisten": asisten_entry.get(),
            "tanggal": tgl_entry.get(),
            "waktu": waktu_entry.get(),
            "cuaca": cuaca_entry.get(),
            "keterangan": ket_textarea.get("1.0", "end-1c")
        }
        # Cek jika ada field yang kosong
        for key, value in data.items():
            if not value:  # Jika ada yang kosong
                messagebox.showwarning("Error", "Pilih data untuk diisi")
                new_window.attributes("-topmost", True)
                new_window.after(10, lambda: new_window.attributes("-topmost", False))  # Mengembalikan normal setelah berada di depan
                return
            

        # Simpan data menggunakan fungsi dari database_manager
        
        database.save_info(data)
        messagebox.showinfo("Success", "Data has been saved")
        new_window.destroy()
    
    
    # judul
    info_label = customtkinter.CTkLabel(new_window, font=font1,  text="INFO MEASUREMENT :", text_color="#fff")
    info_label.place(x=370, y=15)
    
    
    operator_label = customtkinter.CTkLabel(new_window, font=font3,  text="Operator :", text_color="#fff")
    operator_label.place(x=20, y=50)
    operator_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    operator_entry.place(x=150, y=50)
    
    lokasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Lokasi :", text_color="#fff")
    lokasi_label.place(x=20, y=80)
    lokasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    lokasi_entry.place(x=150, y=80)
    
    lintasan_label = customtkinter.CTkLabel(new_window, font=font3,  text="Lintasan :", text_color="#fff")
    lintasan_label.place(x=20, y=110)
    lintasan_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    lintasan_entry.place(x=150, y=110)
    
    pjglintasan_label = customtkinter.CTkLabel(new_window, font=font3,  text="Panjang Lintasan :", text_color="#fff")
    pjglintasan_label.place(x=20, y=140)
    pjglintasan_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    pjglintasan_entry.place(x=150, y=140)
    
    spasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Spasi :", text_color="#fff")
    spasi_label.place(x=195, y=140)
    spasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    spasi_entry.place(x=240, y=140)
    
    jmlchanel_label = customtkinter.CTkLabel(new_window, font=font3,  text="Jumlah Chanel :", text_color="#fff")
    jmlchanel_label.place(x=290, y=140)
    jmlchanel_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    jmlchanel_entry.place(x=383, y=140)
    
    korawal_label = customtkinter.CTkLabel(new_window, font=font3,  text="Kordinat Awal :", text_color="#fff")
    korawal_label.place(x=20, y=170)
    korawal_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    korawal_entry.place(x=150, y=170)
    
    korakhir_label = customtkinter.CTkLabel(new_window, font=font3,  text="Kordinat Akhir :", text_color="#fff")
    korakhir_label.place(x=20, y=200)
    korakhir_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    korakhir_entry.place(x=150, y=200)
    
    elevasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Elevasi :", text_color="#fff")
    elevasi_label.place(x=20, y=230)
    elevasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    elevasi_entry.place(x=150, y=230)
    
    save_entry = customtkinter.CTkButton(new_window, command=save_data, text="Save", font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", cursor="hand2", hover_color="#0C9295", border_width=2, width=100)
    save_entry.place(x=150, y=280)
    
    
    # Membuat sebelah kanan
    asisten_label = customtkinter.CTkLabel(new_window, font=font3,  text="Asisten :", text_color="#fff")
    asisten_label.place(x=450, y=50)
    asisten_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    asisten_entry.place(x=555, y=50)
    
    tgl_label = customtkinter.CTkLabel(new_window, font=font3,  text="Tanggal :", text_color="#fff")
    tgl_label.place(x=450, y=80)
    tgl_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    tgl_entry.place(x=555, y=80)
    
    waktu_label = customtkinter.CTkLabel(new_window, font=font3,  text="Waktu :", text_color="#fff")
    waktu_label.place(x=450, y=110)
    waktu_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    waktu_entry.place(x=555, y=110)
    
    cuaca_label = customtkinter.CTkLabel(new_window, font=font3,  text="Cuaca :", text_color="#fff")
    cuaca_label.place(x=450, y=140)
    cuaca_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    cuaca_entry.place(x=555, y=140)
    
    ket_label = customtkinter.CTkLabel(new_window, font=font3, text="Keterangan:", text_color="#fff")
    ket_label.place(x=450, y=170)
    ket_textarea = customtkinter.CTkTextbox(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, height=90, width=275)
    ket_textarea.place(x=555, y=170)    
    
    
    new_window.attributes("-topmost", True)
    new_window.after(10, lambda: new_window.attributes("-topmost", False))  # Mengembalikan normal setelah berada di depan

info_button = customtkinter.CTkButton(app, text="Info", command=open_info_measurement, font=font1, text_color="#fff", fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=130)
info_button.place(x=20, y=7)

def detail_info():
# Buat jendela baru
    new_window = customtkinter.CTkToplevel(app)
    new_window.title("Info Measurement")
    new_window.geometry("860x400")
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")

    def load_operator_list():
        """Fetch daftar operator dari database."""
        measurements = database.fetch_measurementsInfo()
        operators = [row[1] for row in measurements]  # Ambil kolom operator
        return list(set(operators))  # Hilangkan duplikat

    def load_data_by_operator():
        """Muat data berdasarkan operator yang dipilih."""
        selected_operator = operator_combobox.get()
        measurements = database.fetch_measurementsInfo()

        if not selected_operator:
            messagebox.showwarning("Peringatan", "Silakan pilih operator terlebih dahulu.")
            new_window.attributes("-topmost", True)
            new_window.after(10, lambda: new_window.attributes("-topmost", False))
            return
        # Reset semua entry sebelum memuat data baru
        reset_entries()

        for row in measurements:
            if row[1] == selected_operator:
                # Isi semua entry sesuai data dari database
                idinfo_entry.insert(0, row[0])
                operator_entry.insert(0, row[1])
                lokasi_entry.insert(0, row[2])
                lintasan_entry.insert(0, row[3])
                pjglintasan_entry.insert(0, row[4])
                spasi_entry.insert(0, row[5])
                jmlchanel_entry.insert(0, row[6])
                korawal_entry.insert(0, row[7])
                korakhir_entry.insert(0, row[8])
                elevasi_entry.insert(0, row[9])
                asisten_entry.insert(0, row[10])
                tgl_entry.insert(0, row[11])
                waktu_entry.insert(0, row[12])
                cuaca_entry.insert(0, row[13])
                ket_textarea.insert("1.0", row[14])
                
                idinfo_entry.configure(state="disabled")
            
    def lihat_data():
        """Fungsi untuk memuat data saat tombol diklik."""
        load_data_by_operator()

    def reset_entries():
        """Kosongkan semua entry."""
        entries = [
            idinfo_entry, operator_entry, lokasi_entry, lintasan_entry,
            pjglintasan_entry, spasi_entry, jmlchanel_entry, korawal_entry,
            korakhir_entry, elevasi_entry, asisten_entry, tgl_entry,
            waktu_entry, cuaca_entry
        ]
        for entry in entries:
            entry.delete(0, "end")
        ket_textarea.delete("1.0", "end")

    def save_data():
        """Simpan data baru atau update data ke database."""
        data = {
            "operator": operator_entry.get(),
            "lokasi": lokasi_entry.get(),
            "lintasan": lintasan_entry.get(),
            "panjang_lintasan": pjglintasan_entry.get(),
            "spasi": spasi_entry.get(),
            "jumlah_chanel": jmlchanel_entry.get(),
            "kordinat_awal": korawal_entry.get(),
            "kordinat_akhir": korakhir_entry.get(),
            "elevasi": elevasi_entry.get(),
            "asisten": asisten_entry.get(),
            "tanggal": tgl_entry.get(),
            "waktu": waktu_entry.get(),
            "cuaca": cuaca_entry.get(),
            "keterangan": ket_textarea.get("1.0", "end-1c"),
        }
        
        if not all(data.values()):  # Cek jika ada field kosong
            messagebox.showwarning("Peringatan", "Semua kolom harus diisi sebelum menyimpan data.")
            return

        database.update_info(data)
        messagebox.showinfo("Sukses", "Data berhasil disimpan.")
        new_window.destroy()

    # Label dan input untuk form
    info_label = customtkinter.CTkLabel(new_window, text="DETAIL INFO:", font=font1, text_color="#fff")
    info_label.place(x=370, y=15)

    # Dropdown Pilihan Operator
    operator_label = customtkinter.CTkLabel(new_window, text="Pilih Operator:", font=font3, text_color="#fff")
    operator_label.place(x=540, y=20)
    operator_combobox = customtkinter.CTkComboBox(
        new_window,
        values=load_operator_list(),
        font=font3,
        text_color='#000',
        fg_color="#fff",
        border_color="#0C9295",
        border_width=2,
        width=200
    )
    operator_combobox.place(x=630, y=20)

    # Tombol untuk memuat data
    btn_operator = customtkinter.CTkButton(
        new_window, command=lihat_data,
        text="Klik untuk melihat Data", font=font3, text_color='#000',
        fg_color="#fff", border_color="#0C9295", cursor="hand2",
        hover_color="#0C9295", border_width=2, width=200
    )
    btn_operator.place(x=630, y=50)

    # Entry untuk ID Info
    idinfo_label = customtkinter.CTkLabel(new_window, font=font3, text="ID Info:", text_color="#fff")
    idinfo_label.place(x=20, y=50)
    idinfo_entry = customtkinter.CTkEntry(
        new_window, font=font3, text_color='#000', fg_color="#fff",
        border_color="#0C9295", border_width=2, width=100
    )
    idinfo_entry.place(x=150, y=50)

    operator_label = customtkinter.CTkLabel(new_window, font=font3,  text="Operator :", text_color="#fff")
    operator_label.place(x=20, y=80)
    operator_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    operator_entry.place(x=150, y=80)
    
    lokasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Lokasi :", text_color="#fff")
    lokasi_label.place(x=20, y=110)
    lokasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    lokasi_entry.place(x=150, y=110)
    
    lintasan_label = customtkinter.CTkLabel(new_window, font=font3,  text="Lintasan :", text_color="#fff")
    lintasan_label.place(x=20, y=140)
    lintasan_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    lintasan_entry.place(x=150, y=140)
    
    pjglintasan_label = customtkinter.CTkLabel(new_window, font=font3,  text="Panjang Lintasan :", text_color="#fff")
    pjglintasan_label.place(x=20, y=170)
    pjglintasan_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    pjglintasan_entry.place(x=150, y=170)
    
    spasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Spasi :", text_color="#fff")
    spasi_label.place(x=195, y=170)
    spasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    spasi_entry.place(x=240, y=170)
    
    jmlchanel_label = customtkinter.CTkLabel(new_window, font=font3,  text="Jumlah Chanel :", text_color="#fff")
    jmlchanel_label.place(x=290, y=170)
    jmlchanel_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=40)
    jmlchanel_entry.place(x=383, y=170)
    
    korawal_label = customtkinter.CTkLabel(new_window, font=font3,  text="Kordinat Awal :", text_color="#fff")
    korawal_label.place(x=20, y=200)
    korawal_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    korawal_entry.place(x=150, y=200)
    
    korakhir_label = customtkinter.CTkLabel(new_window, font=font3,  text="Kordinat Akhir :", text_color="#fff")
    korakhir_label.place(x=20, y=230)
    korakhir_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    korakhir_entry.place(x=150, y=230)
    
    elevasi_label = customtkinter.CTkLabel(new_window, font=font3,  text="Elevasi :", text_color="#fff")
    elevasi_label.place(x=20, y=260)
    elevasi_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    elevasi_entry.place(x=150, y=260)
    
    save_entry = customtkinter.CTkButton(new_window, command=save_data, text="Save", font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", cursor="hand2", hover_color="#0C9295", border_width=2, width=100)
    save_entry.place(x=150, y=320)
    
    
    # Membuat sebelah kanan
    asisten_label = customtkinter.CTkLabel(new_window, font=font3,  text="Asisten :", text_color="#fff")
    asisten_label.place(x=450, y=80)
    asisten_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    asisten_entry.place(x=555, y=80)
    
    tgl_label = customtkinter.CTkLabel(new_window, font=font3,  text="Tanggal :", text_color="#fff")
    tgl_label.place(x=450, y=110)
    tgl_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    tgl_entry.place(x=555, y=110)
    
    waktu_label = customtkinter.CTkLabel(new_window, font=font3,  text="Waktu :", text_color="#fff")
    waktu_label.place(x=450, y=140)
    waktu_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    waktu_entry.place(x=555, y=140)
    
    cuaca_label = customtkinter.CTkLabel(new_window, font=font3,  text="Cuaca :", text_color="#fff")
    cuaca_label.place(x=450, y=170)
    cuaca_entry = customtkinter.CTkEntry(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=275)
    cuaca_entry.place(x=555, y=170)
    
    ket_label = customtkinter.CTkLabel(new_window, font=font3, text="Keterangan:", text_color="#fff")
    ket_label.place(x=450, y=200)
    ket_textarea = customtkinter.CTkTextbox(new_window, font=font3, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, height=90, width=275)
    ket_textarea.place(x=555, y=200)    
    
    new_window.attributes("-topmost", True)
    new_window.after(10, lambda: new_window.attributes("-topmost", False))


def close():
    app.destroy()

detail_info = customtkinter.CTkButton(app, text="Detail", command=detail_info, font=font1, text_color="#fff", fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=130)
detail_info.place(x=150, y=7)

time_label = customtkinter.CTkLabel(app, font=font1, text="", text_color="#fff",  width=300, height=30)
time_label.place(x=380, y=7)

# id_label = customtkinter.CTkLabel(app, font=font1,  text="No:", text_color="#fff")
# id_label.place(x=20, y=50)

id_entry = customtkinter.CTkEntry(app, font=font1, text_color='#1f1f1f', fg_color="#1f1f1f", border_color="#1f1f1f", border_width=2, width=180)
id_entry.place(x=100, y=50)

close = customtkinter.CTkButton(app, font=font1, text="Close", text_color='#1f1f1f', fg_color="#1f1f1f", border_color="#1f1f1f", border_width=2, width=180)
close.place(x=100, y=50)


volt_label = customtkinter.CTkLabel(app, font=font1,  text="Volt:", text_color="#fff")
volt_label.place(x=20, y=100)

volt_entry = customtkinter.CTkEntry(app, font=font1, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=180)
volt_entry.place(x=100, y=100)

ampere_label = customtkinter.CTkLabel(app, font=font1,  text="Ampere:", text_color="#fff")
ampere_label.place(x=20, y=150)

ampere_entry = customtkinter.CTkEntry(app, font=font1, text_color='#000', fg_color="#fff", border_color="#0C9295", border_width=2, width=180)
ampere_entry.place(x=100, y=150)

csv_button = customtkinter.CTkButton(app, command=export_to_csv, font=font1, text_color="#fff", text="Export to CSV", fg_color="#4F75FF", hover_color="#6439FF", cursor="hand2", corner_radius=15,  width=260)
csv_button.place(x=20, y=210)

pdf_button = customtkinter.CTkButton(app, command=export_to_pdf, font=font1, text_color="#fff", text="Export to PDF", fg_color="#E40404", hover_color="#AE0000", cursor="hand2", corner_radius=15,  width=260)
pdf_button.place(x=20, y=260)

excel_button = customtkinter.CTkButton(app, command=export_to_excel, font=font1, text_color="#fff", text="Export to Excel", fg_color="#054312", hover_color="#00850B", cursor="hand2", corner_radius=15,  width=260)
excel_button.place(x=20, y=310)


# Button start measurement
start_button = customtkinter.CTkButton(app, command=start_measurement, font=font1, text_color="#fff", text="Start Measurement", fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=260)
start_button.place(x=20, y=360)

# Pengaturan tombol dan entri input
prev_button = customtkinter.CTkButton(app, command=lambda: update_target_input(-1), text="Prev", font=font1, fg_color="#4F75FF", hover_color="#6439FF", corner_radius=15, width=70)
prev_button.place(x=20, y=420)

target_input_entry = customtkinter.CTkEntry(app, textvariable=target_input, font=font1, fg_color="#161C25", border_color="#F15704", border_width=2, corner_radius=15, width=100)
target_input_entry.place(x=100, y=420)

next_button = customtkinter.CTkButton(app, command=lambda: update_target_input(1), text="Next", font=font1, fg_color="#4F75FF", hover_color="#6439FF", corner_radius=15, width=70)
next_button.place(x=210, y=420)

auto_button = customtkinter.CTkButton(app, command=auto_measurement, font=font1, text_color="#fff", text="Automatic Measurement", fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=260)
auto_button.place(x=300, y=420)

manual_button = customtkinter.CTkButton(app, command=manual_measurement, font=font1, text_color="#fff", text="Manual Measurement", fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=260)
manual_button.place(x=580, y=420)



update_button = customtkinter.CTkButton(app, command=update, font=font1, text_color="#fff", text="Update Measurement",  fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=260)
update_button.place(x=300, y=360)

delete_button = customtkinter.CTkButton(app, command=delete, font=font1, text_color="#fff", text="Delete Measurement",  fg_color="#161C25", hover_color="#FF5002", border_color="#F15704", border_width=2, cursor="hand2", corner_radius=15, width=260)
delete_button.place(x=580, y=360)

# Style untuk Treeview
style = ttk.Style(app)
style.theme_use("clam")
style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))
style.map("Treeview", background=[("selected", "#1A8F2D")])

# Menggunakan Notebook untuk menambahkan beberapa sheet
notebook = ttk.Notebook(app)
notebook.place(x=300, y=40, width=530, height=270)

# Fungsi untuk membuat Treeview pada tiap sheet
def create_treeview(parent):
    tree = ttk.Treeview(parent, height=16)
    tree["columns"] = ("ID","Volt", "Ampere", "Time")

    tree.column("#0", width=0, stretch=tk.NO)
    tree.column("ID", anchor=tk.CENTER, width=120)
    tree.column("Volt", anchor=tk.CENTER, width=120)
    tree.column("Ampere", anchor=tk.CENTER, width=120)
    tree.column("Time", anchor=tk.CENTER, width=170)

    tree.heading("ID", text="ID")
    tree.heading("Volt", text="Volt")
    tree.heading("Ampere", text="Ampere")
    tree.heading("Time", text="Time")

    tree.pack(expand=True, fill="both")
    tree.bind("<ButtonRelease-1>", display_data)
    return tree

# Sheet pertama
sheet1 = ttk.Frame(notebook)
notebook.add(sheet1, text="Sheet 1")
tree = create_treeview(sheet1)

# Fungsi untuk menambahkan sheet baru saat tombol "+" diklik
def add_new_sheet():
    sheet_num = len(notebook.tabs()) + 1
    new_sheet = ttk.Frame(notebook)
    notebook.add(new_sheet, text=f"Sheet {sheet_num}")
    create_treeview(new_sheet)

# Tombol tambah sheet
plus_btn = customtkinter.CTkLabel(app, text="+", font=("Helvetica", 14, "bold"), text_color="#fff", fg_color="#E40404", cursor="hand2", width=50)
plus_btn.place(x=400, y=310)
plus_btn.bind("<Button-1>", lambda e: add_new_sheet())

# Label navigasi sheet, ganti saat tab di klik
def on_tab_change(event):
    current_tab = notebook.tab(notebook.select(), "text")
    table_label.configure(text=current_tab)

table_label = customtkinter.CTkLabel(app, text="Sheet 1", font=("Helvetica", 14), text_color="#fff", fg_color="#4F75FF",  cursor="hand2", width=100)
table_label.place(x=300, y=310)

# Menghubungkan perubahan tab dengan update label
notebook.bind("<<NotebookTabChanged>>", on_tab_change)


show_current_time()
# add_to_treeview()

app.mainloop()